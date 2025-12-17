#!/usr/bin/env python3
"""
step5f_make_labeling_xlsx.py

Convert an error_sheet CSV into an Excel .xlsx that is easier to label:
- Adds computed columns: f1_baseline, f1_improved, delta_f1, winloss
- Adds dropdown validations for label columns
- Optionally auto-fills obvious retrieval errors when labels are blank

Run:
  python step5f_make_labeling_xlsx.py ^
    --in_csv outputs/predictions/bart_debug/error_sheet_val_trigger_80.csv ^
    --out_xlsx outputs/predictions/bart_debug/error_sheet_val_trigger_80_labeled.xlsx
"""

from __future__ import annotations
import argparse
import csv
import re
import string
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


#EM/F1 normalization
_ARTICLES = re.compile(r"\b(a|an|the)\b", re.IGNORECASE)
_PUNCT_TABLE = str.maketrans("", "", string.punctuation)
_WS = re.compile(r"\s+")


def normalize_answer(s: str) -> str:
    s = (s or "").lower()
    s = s.translate(_PUNCT_TABLE)
    s = _ARTICLES.sub(" ", s)
    s = _WS.sub(" ", s).strip()
    return s


def f1_score(pred: str, ref: str) -> float:
    pt = normalize_answer(pred).split()
    rt = normalize_answer(ref).split()
    if not pt and not rt:
        return 1.0
    if not pt or not rt:
        return 0.0
    common = Counter(pt) & Counter(rt)
    num_same = sum(common.values())
    if num_same == 0:
        return 0.0
    precision = num_same / len(pt)
    recall = num_same / len(rt)
    return (2 * precision * recall) / (precision + recall)


def read_csv(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    with open(path, "r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        header = list(r.fieldnames or [])
        rows = [dict(row) for row in r]
    return header, rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in_csv", type=str, required=True)
    ap.add_argument("--out_xlsx", type=str, required=True)
    ap.add_argument("--autofill_obvious", action="store_true", help="Autofill clear retrieval misses when labels are blank.")
    args = ap.parse_args()

    in_csv = Path(args.in_csv)
    out_xlsx = Path(args.out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    header, rows = read_csv(in_csv)
    if not rows:
        raise RuntimeError("CSV has 0 data rows.")

    # Auto-detect baseline/improved columns from prediction
    pred_cols = [c for c in header if c.startswith("prediction__")]
    if len(pred_cols) < 2:
        raise RuntimeError(f"Expected at least 2 prediction__* columns, found {pred_cols}")

    base_pred_col = pred_cols[0]
    imp_pred_col = pred_cols[1]
    base_suffix = base_pred_col[len("prediction__"):]
    imp_suffix = imp_pred_col[len("prediction__"):]

    # Related columns 
    base_doc_col = f"doc_hit__{base_suffix}"
    base_span_col = f"span_hit__{base_suffix}"
    imp_doc_col = f"doc_hit__{imp_suffix}"
    imp_span_col = f"span_hit__{imp_suffix}"

    required = ["reference", "user_turn", "label_error_type", "label_who_failed", "label_notes"]
    for req in required:
        if req not in header:
            raise RuntimeError(f"Missing required column in CSV: {req}")

    # Add computed columns at the end
    computed_cols = ["f1_baseline", "f1_improved", "delta_f1", "winloss"]
    final_header = header + [c for c in computed_cols if c not in header]

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "error_sheet"

    # Hidden validation sheet
    vs = wb.create_sheet("validations")
    vs.sheet_state = "hidden"

    who_failed_vals = ["retrieval", "rewrite", "generation", "evaluation", "data", "none"]
    err_type_vals = [
        "RET_DOC_MISS", "RET_SPAN_MISS", "RET_OK",
        "RW_WRONG_ANTECEDENT", "RW_OVERCONSTRAIN", "RW_UNDERINFORM", "RW_TRIGGER_ERROR",
        "GEN_IGNORES_EVIDENCE", "GEN_UNGROUNDED", "GEN_INCOMPLETE", "GEN_WRONG_ENTITY",
        "EVAL_PARAPHRASE", "REF_AMBIGUOUS",
    ]

    # Write validation lists
    for i, v in enumerate(who_failed_vals, start=1):
        vs.cell(row=i, column=1).value = v
    for i, v in enumerate(err_type_vals, start=1):
        vs.cell(row=i, column=2).value = v

    who_range = f"validations!$A$1:$A${len(who_failed_vals)}"
    err_range = f"validations!$B$1:$B${len(err_type_vals)}"

    # Header row
    bold = Font(bold=True)
    for j, col in enumerate(final_header, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Precompute column indices for label cols
    col_index = {c: i + 1 for i, c in enumerate(final_header)}
    who_col = col_index["label_who_failed"]
    err_col = col_index["label_error_type"]

    # Write rows
    wrap_cols = [
        "history",
        "evidence_preview__" + base_suffix,
        "evidence_preview__" + imp_suffix,
        "label_notes",
    ]
    wrap_set = set([c for c in wrap_cols if c in col_index])

    for i, row in enumerate(rows, start=2):
        ref = row.get("reference", "")
        base_pred = row.get(base_pred_col, "")
        imp_pred = row.get(imp_pred_col, "")

        f1_b = f1_score(base_pred, ref)
        f1_i = f1_score(imp_pred, ref)
        delta = f1_i - f1_b
        if delta > 1e-12:
            wl = "win"
        elif delta < -1e-12:
            wl = "loss"
        else:
            wl = "tie"

        # Optional: autofill obvious retrieval issues if label empty
        if args.autofill_obvious and not (row.get("label_who_failed", "").strip() or row.get("label_error_type", "").strip()):
            # Only do this when BOTH systems fail retrieval in the same obvious way
            b_doc = int(row.get(base_doc_col, "0") or "0") if base_doc_col in row else 0
            i_doc = int(row.get(imp_doc_col, "0") or "0") if imp_doc_col in row else 0
            b_span = int(row.get(base_span_col, "0") or "0") if base_span_col in row else 0
            i_span = int(row.get(imp_span_col, "0") or "0") if imp_span_col in row else 0

            if b_doc == 0 and i_doc == 0:
                row["label_who_failed"] = "retrieval"
                row["label_error_type"] = "RET_DOC_MISS"
            elif b_doc == 1 and i_doc == 1 and b_span == 0 and i_span == 0:
                row["label_who_failed"] = "retrieval"
                row["label_error_type"] = "RET_SPAN_MISS"

        # Fill row values
        for j, col in enumerate(final_header, start=1):
            if col == "f1_baseline":
                val = round(f1_b, 4)
            elif col == "f1_improved":
                val = round(f1_i, 4)
            elif col == "delta_f1":
                val = round(delta, 4)
            elif col == "winloss":
                val = wl
            else:
                val = row.get(col, "")
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(wrap_text=(col in wrap_set), vertical="top")

    # Freeze header + enable filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths (safe defaults)
    default_width = 18
    wide_cols = {
        "example_id": 45,
        "user_turn": 35,
        "reference": 45,
        "history": 60,
        "label_notes": 40,
    }
    for col, idx in col_index.items():
        width = wide_cols.get(col, default_width)
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Add dropdown validations
    dv_who = DataValidation(type="list", formula1=f"={who_range}", allow_blank=True)
    dv_err = DataValidation(type="list", formula1=f"={err_range}", allow_blank=True)
    ws.add_data_validation(dv_who)
    ws.add_data_validation(dv_err)

    # Apply to full column range (row 2..last)
    last_row = 1 + len(rows)
    dv_who.add(f"{get_column_letter(who_col)}2:{get_column_letter(who_col)}{last_row}")
    dv_err.add(f"{get_column_letter(err_col)}2:{get_column_letter(err_col)}{last_row}")

    wb.save(out_xlsx)
    print("Wrote:", out_xlsx)
    print("Detected baseline:", base_pred_col)
    print("Detected improved:", imp_pred_col)
    print("Added computed columns:", computed_cols)


if __name__ == "__main__":
    main()
