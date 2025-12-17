#!/usr/bin/env python3
"""
step5g_summarize_labels.py

Summarize labeled error analysis sheet (CSV or XLSX).
Outputs counts for label_who_failed and label_error_type, plus a small table.

Run:
  python step5g_summarize_labels.py --in_file outputs/predictions/bart_debug/error_sheet_val_trigger_80_labeled.xlsx
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def read_xlsx(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(path)
    ws = wb["error_sheet"] if "error_sheet" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x) for x in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        for k, v in zip(header, r):
            d[k] = "" if v is None else str(v)
        out.append(d)
    return out


def read_csv(path: Path) -> List[Dict[str, str]]:
    with open(path, "r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        return [dict(row) for row in r]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in_file", type=str, required=True)
    args = ap.parse_args()

    path = Path(args.in_file)
    if not path.exists():
        raise FileNotFoundError(path)

    if path.suffix.lower() == ".xlsx":
        rows = read_xlsx(path)
    else:
        rows = read_csv(path)

    who = Counter()
    et = Counter()
    labeled = 0

    for r in rows:
        w = (r.get("label_who_failed", "") or "").strip()
        e = (r.get("label_error_type", "") or "").strip()
        if w or e:
            labeled += 1
        if w:
            who[w] += 1
        if e:
            et[e] += 1

    print(f"Rows: {len(rows)} | Labeled rows: {labeled}")
    print("\nlabel_who_failed counts:")
    for k, v in who.most_common():
        print(f"  {k}: {v}")

    print("\nlabel_error_type counts:")
    for k, v in et.most_common():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
